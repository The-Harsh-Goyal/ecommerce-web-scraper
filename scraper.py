from time import *
from random import *
from requests import *
from bs4 import BeautifulSoup as BS
import pandas as pd
from datetime import *
from os import *
from openpyxl import *
from openpyxl.styles import *

# --- HEADERS (shared across all scraping calls) ---
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

DELAY_RANGE = (2, 5)


def extract_with_fallbacks(soup_or_element, selectors, attr=None):
    """
    Try a list of CSS selectors in order.
    - If attr is None: return cleaned text.
    - If attr is provided: return that attribute value.
    - If nothing matches: return None.
    """
    if soup_or_element is None:
        return None

    for selector in selectors:
        try:
            el = soup_or_element.select_one(selector)
        except Exception:
            el = None

        if el is None:
            continue

        if attr is not None:
            value = el.get(attr)
            if value:
                return value.strip()
        else:
            text = el.get_text(strip=True)
            if text:
                return text

    return None


def fetch_page(url: str):
    """
    Fetch page HTML from a URL.
    Returns: (html_string, error_message)
    """
    try:
        resp = get(url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        return resp.text, None
    except Exception as e:
        return None, str(e)


def parse_products(html: str):
    """
    Parse e-commerce products from HTML.
    Returns: list of dicts with product data
    """
    soup = BS(html, "html.parser")

    # FIXED: Product block selectors – ordered by specificity & compatibility
    product_block_selectors = [
        ".product",                 # web-scraping.dev ✓ WORKS (keep this primary!)
        ".row.product",             # Alternative pattern
        ".product-card",            # Generic e-comm
        ".product-tile",            # Generic e-comm
        ".product-item",            # Generic e-comm
        ".product-listing",         # Generic e-comm
        ".plp-card",                # Generic e-comm
        # NEW: Flipkart selectors
        "div._4ddWXP",              # Flipkart product container
        # NEW: Walmart selectors
        "div[data-automation-id*='product']",  # Walmart product
        # NEW: eBay selectors
        "div.s-item",               # eBay product item
    ]

    product_blocks = []
    for sel in product_block_selectors:
        found = soup.select(sel)
        if found:
            product_blocks = found
            print(f"[DEBUG] Found {len(found)} products with selector: {sel}")
            break

    products = []

    # ===== TITLE SELECTORS =====
    TITLE_SELECTORS = [
        ".product-name",
        "h2.product-name",
        "h3.product-name",
        "h2.product-title",
        "h3.product-title",
        ".product-title",
        "h2",
        "h3",
    ]

    # ===== PRICE SELECTORS =====
    # MERGED: Current + Your new ones (no duplicates)
    PRICE_SELECTORS = [
        ".product-price",
        ".price",
        "span.price",
        ".sale-price",
        "[class*='price']",
        # NEW ADDITIONS:
        ".current-price",           # Generic current price
        ".item-price",              # Generic item price
        # NEW: Flipkart price selectors
        "div._30jeq3",              # Flipkart primary price
        "span._16Jk6d",             # Flipkart alternate price
        # NEW: Walmart price selectors
        "span[data-automation-id*='price']",  # Walmart price
        # NEW: eBay price selectors
        "span.s-item-price",        # eBay item price
    ]

    # ===== AVAILABILITY SELECTORS =====
    # MERGED: Removed duplicate ".in-stock" (appears only once now)
    AVAILABILITY_SELECTORS = [
        ".availability",
        ".stock-status",
        ".in-stock",                # Covers both generic and Flipkart
        ".stock",
        "[class*='availability']",
        "[class*='stock']",
        # NEW ADDITIONS:
        ".stock-info",              # Generic stock info
        # NEW: eBay availability
        "span.SECONDARY_INFO",      # eBay secondary info
    ]

    # ===== LINK SELECTORS =====
    # MERGED: Removed duplicate "a" (only at end), added a[href*='item']
    LINK_SELECTORS = [
        "a.product-link",
        "a.product-name",
        ".product a",
        "a[href*='product']",
        # NEW ADDITION:
        "a[href*='item']",          # Generic item links
        "a",                        # Fallback (last option)
    ]

    # ===== NEW: IMAGE/THUMBNAIL SELECTORS =====
    IMAGE_SELECTORS = [
        "img.product-image",
        "img.product-thumbnail",
        "img.product-thumb",
        ".product-image img",
        "img[src*='product']",
        "img[alt*='product']",
    ]

    # ===== NEW: SELLER/VENDOR SELECTORS =====
    SELLER_SELECTORS = [
        ".seller-name",
        ".vendor-name",
        ".store-name",
        "[class*='seller']",
        "[class*='vendor']",
        "[class*='store']",
    ]

    # ===== NEW: RATING/REVIEW SELECTORS =====
    RATING_SELECTORS = [
        ".product-rating",
        ".rating",
        ".star-rating",
        ".review-score",
        "[class*='rating']",
        "[class*='stars']",
    ]

    # ===== NEW: UNITS SOLD / POPULARITY SELECTORS =====
    SOLD_COUNT_SELECTORS = [
        ".sold-count",
        ".units-sold",
        "[class*='sold']",
        "[class*='purchased']",
        ".popularity",
        ".buy-count",
    ]

    # ===== NEW: CONDITION SELECTORS =====
    CONDITION_SELECTORS = [
        ".condition",
        ".product-condition",
        "[class*='condition']",
        "[class*='refurbish']",
    ]

    # ===== NEW: DELIVERY COST SELECTORS =====
    # FIXED: Removed [class*='delivery'] and [class*='shipping']
    # (moved to DELIVERY_INFO_SELECTORS to avoid duplication)
    DELIVERY_SELECTORS = [
        ".delivery-cost",
        ".shipping-cost",
        ".shipping-price",
    ]

    # ===== NEW: SELLER RATING SELECTORS =====
    SELLER_RATING_SELECTORS = [
        ".seller-rating",
        ".seller-score",
        ".store-rating",
        "[class*='seller-rating']",
        "[class*='positive']",
    ]

    # ===== NEW: ORIGINAL PRICE SELECTORS =====
    ORIGINAL_PRICE_SELECTORS = [
        ".original-price",
        ".old-price",
        ".rrp",
        ".mrp-price",
        "[class*='original-price']",
        "[class*='mrp']",
    ]

    # ===== NEW: DISCOUNT PERCENTAGE SELECTORS =====
    DISCOUNT_PERCENTAGE_SELECTORS = [
        ".discount-percentage",
        ".discount-badge",
        ".percent-off",
        ".savings",
        "[class*='discount-percent']",
        "[class*='off']",
    ]

    # ===== NEW: STOCK COUNT SELECTORS =====
    STOCK_COUNT_SELECTORS = [
        ".stock-count",
        ".quantity-left",
        ".inventory-count",
        ".units-left",
        "[class*='stock-count']",
        "[data-stock]",
    ]

    # ===== NEW: CATEGORY SELECTORS =====
    CATEGORY_SELECTORS = [
        ".breadcrumb",
        ".product-category",
        ".category-tag",
        "[class*='category']",
        "[class*='breadcrumb']",
    ]

    # ===== NEW: MERCHANT SELECTORS =====
    MERCHANT_SELECTORS = [
        ".seller-info",
        ".merchant-name",
        ".store-badge",
        "[class*='seller-badge']",
        "[class*='merchant']",
    ]

    # ===== NEW: DELIVERY INFO SELECTORS =====
    # FIXED: Now includes [class*='delivery'] and [class*='shipping']
    # (removed from DELIVERY_SELECTORS to avoid duplication)
    DELIVERY_INFO_SELECTORS = [
        ".delivery-info",
        ".shipping-info",
        ".delivery-date",
        "[class*='delivery']",      # Moved here (was in DELIVERY_SELECTORS)
        "[class*='shipping-date']",
        "[class*='shipping']",      # Moved here (was in DELIVERY_SELECTORS)
    ]

    # ===== NEW: STOCK STATUS SELECTORS =====
    STOCK_STATUS_SELECTORS = [
        ".out-of-stock",
        ".stock-status",
        "[data-availability]",
        "[class*='low-stock']",
        "[class*='stock-status']",
    ]

    # ===== NEW: BADGE SELECTORS =====
    BADGE_SELECTORS = [
        ".badge",
        ".product-badge",
        ".hot-deal",
        ".sale-badge",
        ".new-product",
        ".best-seller",
        "[class*='badge']",
        "[class*='label']",
    ]

    # ===== NEW: SKU SELECTORS =====
    SKU_SELECTORS = [
        ".product-sku",
        ".sku",
        "[data-product-id]",
        "[data-sku]",
        "[class*='sku']",
    ]

    # ===== NEW: SPECS SELECTORS =====
    SPECS_SELECTORS = [
        ".specs",
        ".highlights",
        ".product-specs",
        ".feature-list",
        "[class*='spec']",
    ]

    # ===== NEW: WARRANTY SELECTORS =====
    WARRANTY_SELECTORS = [
        ".warranty",
        ".guarantee",
        ".warranty-text",
        "[class*='warranty']",
        "[class*='guarantee']",
    ]

    # ===== EXTRACT PRODUCTS =====
    for block in product_blocks:
        title = extract_with_fallbacks(block, TITLE_SELECTORS)
        price = extract_with_fallbacks(block, PRICE_SELECTORS)
        availability = extract_with_fallbacks(block, AVAILABILITY_SELECTORS)
        product_link = extract_with_fallbacks(block, LINK_SELECTORS, attr="href")
        image_url = extract_with_fallbacks(block, IMAGE_SELECTORS, attr="src")
        seller = extract_with_fallbacks(block, SELLER_SELECTORS)
        rating = extract_with_fallbacks(block, RATING_SELECTORS)
        sold_count = extract_with_fallbacks(block, SOLD_COUNT_SELECTORS)
        condition = extract_with_fallbacks(block, CONDITION_SELECTORS)
        delivery_cost = extract_with_fallbacks(block, DELIVERY_SELECTORS)
        seller_rating = extract_with_fallbacks(block, SELLER_RATING_SELECTORS)
        original_price = extract_with_fallbacks(block, ORIGINAL_PRICE_SELECTORS)
        discount_in_percentage = extract_with_fallbacks(block, DISCOUNT_PERCENTAGE_SELECTORS)
        stock_qty = extract_with_fallbacks(block, STOCK_COUNT_SELECTORS)
        product_category = extract_with_fallbacks(block, CATEGORY_SELECTORS)
        seller_info = extract_with_fallbacks(block, MERCHANT_SELECTORS)
        delivery_info = extract_with_fallbacks(block, DELIVERY_INFO_SELECTORS)
        stock_status = extract_with_fallbacks(block, STOCK_STATUS_SELECTORS)
        badge = extract_with_fallbacks(block, BADGE_SELECTORS)
        product_code = extract_with_fallbacks(block, SKU_SELECTORS)
        specs = extract_with_fallbacks(block, SPECS_SELECTORS)
        warranty = extract_with_fallbacks(block, WARRANTY_SELECTORS)

        # Make relative links absolute
        if product_link and product_link.startswith("/"):
            # Detect domain from context
            if "web-scraping.dev" in str(block):
                product_link = "https://www.web-scraping.dev" + product_link
            elif "flipkart" in str(block).lower():
                product_link = "https://www.flipkart.com" + product_link
            elif "walmart" in str(block).lower():
                product_link = "https://www.walmart.com" + product_link
            elif "ebay" in str(block).lower():
                product_link = "https://www.ebay.com" + product_link

        if image_url and image_url.startswith("/"):
            # Similar domain detection for images
            if "web-scraping.dev" in str(block):
                image_url = "https://www.web-scraping.dev" + image_url
            elif "flipkart" in str(block).lower():
                image_url = "https://www.flipkart.com" + image_url
            elif "walmart" in str(block).lower():
                image_url = "https://www.walmart.com" + image_url
            elif "ebay" in str(block).lower():
                image_url = "https://www.ebay.com" + image_url

        products.append(
            {
                "product_name": title,
                "price": price,
                "availability": availability,
                "product_url": product_link,
                "image_url": image_url,
                "seller": seller,
                "rating": rating,
                "units_sold": sold_count,
                "condition": condition,
                "delivery_cost": delivery_cost,
                "seller_rating": seller_rating,
                "original_price": original_price,
                "discount_in_percentage": discount_in_percentage,
                "stock_qty": stock_qty,
                "product_category": product_category,
                "seller_info": seller_info,
                "delivery_info": delivery_info,
                "stock_status": stock_status,
                "badge": badge,
                "product_code": product_code,
                "specs": specs,
                "scraped_date": datetime.now().strftime("%Y-%m-%d"),
                "scraped_time": datetime.now().strftime("%H:%M:%S"),
            }
        )

    return products

# ========== NEW: EXCEL & CSV EXPORT FUNCTIONS ==========

def save_to_csv_and_excel(products, csv_filename: str = None, excel_filename: str = None):
    """
    Save products to both CSV and Excel files with automatic date-based naming.
    
    Features:
    - Automatic filename generation with today's date (yyyyMMdd format)
    - Excel file with formatted headers, auto-sized columns, and borders
    - CSV file for compatibility
    - Creates 'scraped_data' directory if it doesn't exist
    
    Args:
        products (list): List of product dictionaries
        csv_filename (str): Optional custom CSV filename (default: auto-generated)
        excel_filename (str): Optional custom Excel filename (default: auto-generated)
    
    Returns:
        (success: bool, csv_path: str or None, excel_path: str or None, message: str)
    """
    if not products:
        return False, None, None, "❌ No products to save."

    try:
        # Create output directory
        makedirs('scraped_data', exist_ok=True)
        
        # Generate date-based filenames if not provided
        today_date = datetime.now().strftime("%Y%m%d")  # Format: 20251225
        
        if csv_filename is None:
            csv_filename = f"products_{today_date}.csv"
        if excel_filename is None:
            excel_filename = f"products_{today_date}.xlsx"
        
        csv_path = path.join('scraped_data', csv_filename)
        excel_path = path.join('scraped_data', excel_filename)
        
        # ===== SAVE TO CSV =====
        df = pd.DataFrame(products)
        df.to_csv(csv_path, index=False, encoding="utf-8")
        print(f"✅ CSV saved: {csv_path}")
        
        # ===== SAVE TO EXCEL WITH FORMATTING =====
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        
        # Define styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add headers
        headers = df.columns.tolist()
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border_style
        
        # Add data rows
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border_style
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save Excel file
        wb.save(excel_path)
        print(f"✅ Excel saved: {excel_path}")
        
        return True, csv_path, excel_path, f"✅ Saved {len(df)} products to both CSV and Excel"
    
    except Exception as e:
        return False, None, None, f"❌ Error saving files: {str(e)}"


def save_to_csv(products, filename: str):
    """
    Legacy function - keeps backward compatibility.
    Now calls save_to_csv_and_excel().
    """
    success, csv_path, excel_path, message = save_to_csv_and_excel(products, filename)
    return success, message


# ========== MAIN SCRAPING FUNCTION ==========

def scrape_and_save(url: str, custom_csv_filename: str = None, custom_excel_filename: str = None):
    """
    Main function: fetch -> parse -> save to CSV & Excel.
    
    Automatically generates filenames with today's date (yyyyMMdd format).
    
    Args:
        url (str): URL to scrape
        custom_csv_filename (str): Optional custom CSV filename
        custom_excel_filename (str): Optional custom Excel filename
    
    Returns:
        (success: bool, message: str, dataframe: pd.DataFrame or None, product_count: int)
    """
    print(f"\n{'='*60}")
    print(f"🚀 Starting scrape for: {url}")
    print(f"{'='*60}\n")
    
    # Step 1: Fetch
    html, fetch_error = fetch_page(url)
    if fetch_error:
        error_msg = f"❌ Failed to fetch: {fetch_error}"
        print(error_msg)
        return False, error_msg, None, 0

    # Step 2: Polite delay
    delay = uniform(*DELAY_RANGE)
    print(f"⏳ Polite delay: {delay:.2f} seconds")
    sleep(delay)

    # Step 3: Parse
    print("🔍 Parsing products...")
    products = parse_products(html)
    
    if not products:
        error_msg = "❌ No products found on the page."
        print(error_msg)
        return False, error_msg, None, 0

    print(f"✅ Found {len(products)} products\n")

    # Step 4: Save to CSV and Excel
    print("💾 Saving data...")
    success, csv_path, excel_path, save_msg = save_to_csv_and_excel(
        products, 
        custom_csv_filename, 
        custom_excel_filename
    )
    
    if success:
        df = pd.DataFrame(products)
        print(f"\n{save_msg}")
        print(f"   📁 CSV: {csv_path}")
        print(f"   📁 Excel: {excel_path}")
        print(f"   📊 Total records: {len(df)}")
        print(f"{'='*60}\n")
        return True, save_msg, df, len(products)
    else:
        print(f"\n{save_msg}")
        return False, save_msg, None, 0

# --- CLI mode (if running scraper.py directly) ---
if __name__ == "__main__":
    URL = "https://www.web-scraping.dev/products"
    
    # Run scraper (filenames auto-generated with today's date)
    success, message, df, count = scrape_and_save(URL)
    
    if df is not None:
        print("\n📋 First 5 products:")
        print(df.head())

